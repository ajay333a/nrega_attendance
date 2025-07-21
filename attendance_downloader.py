import io
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font

STARTING_URL = "https://mnregaweb4.nic.in/nregaarch/View_NMMS_atten_date_dtl_rpt.aspx?page=&short_name=KN&state_name=KARNATAKA&state_code=15&district_name=BALLARI&district_code=1505&block_name=SIRUGUPPA&block_code=1505007&"
DEFAULT_DISTRICT = "Ballari"
DEFAULT_TALUK = "Siruguppa"


def get_attendance_data(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching attendance data: {e}")
        return None, None, None, None
    soup = BeautifulSoup(response.content, 'html.parser')
    # Extract work name
    work_name = None
    for b in soup.find_all('b'):
        if b.text.strip().startswith('Work Name'):
            next_text = b.next_sibling
            if next_text:
                work_name = str(next_text).strip(' :\u00a0-')
            break
    if not work_name:
        work_name_elem = soup.find(id="ContentPlaceHolder1_lbl_dtl")
        if work_name_elem:
            work_name = work_name_elem.text.strip()
    # Attendance Table
    attendance_data = []
    tables = soup.find_all('table')
    if not tables:
        print("No tables found on the page.")
        return None, None, work_name, None
    attendance_table = tables[-1]
    rows = attendance_table.find_all('tr')
    header_cells = [th.text.strip() for th in rows[0].find_all(['th', 'td'])]
    col_map = {name: idx for idx, name in enumerate(header_cells)}
    for row in rows[1:]:
        cols = row.find_all('td')
        if cols and any(c.get_text(strip=True) for c in cols):
            name_td = ''
            for td in cols:
                span = td.find('span', id=lambda x: x and 'lbl_workerName_' in x)
                if span:
                    name_td = span.get_text(strip=True)
                    break
            extracted = [
                cols[col_map.get('S.No', -1)].get_text(strip=True) if 'S.No' in col_map else '',
                cols[col_map.get('Job Card No', -1)].get_text(strip=True) if 'Job Card No' in col_map else '',
                name_td,
                cols[col_map.get('Attendance Date', -1)].get_text(strip=True) if 'Attendance Date' in col_map else '',
                cols[col_map.get('Present/Absent', -1)].get_text(strip=True) if 'Present/Absent' in col_map else ''
            ]
            attendance_data.append(extracted)
    photo_url = None
    img_link = soup.find('a', text='Click here for large image')
    if img_link and img_link.has_attr('href'):
        photo_url = img_link['href']
    return attendance_data, photo_url, work_name, header_cells


def download_photo(url):
    if not url:
        return None
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        img_bytes = io.BytesIO(response.content)
        return img_bytes
    except requests.exceptions.RequestException as e:
        print(f"Error downloading photo: {e}")
        return None


def write_attendance_excel(attendance_records, work_code, work_name, panchayat_name, file_base):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Data'
    ws.append([f'Work Code: {work_code}'])
    ws.append([f'Work Name: {work_name if work_name else ""}'])
    ws.append([f'District: {DEFAULT_DISTRICT}'])
    ws.append([f'Taluk/Block: {DEFAULT_TALUK}'])
    ws.append([f'Panchayath Name: {panchayat_name}'])
    ws.append([])
    for record in attendance_records:
        ws.append([record['muster_roll_no']] + record['row'])
    print(f'Saved attendance_data_{file_base}.xlsx')
    return wb


def write_images_excel(image_records, work_code, work_name, panchayat_name, file_base):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Images'
    ws.append([f'Work Code: {work_code}'])
    ws.append([f'Work Name: {work_name if work_name else ""}'])
    ws.append([f'District: {DEFAULT_DISTRICT}'])
    ws.append([f'Taluk/Block: {DEFAULT_TALUK}'])
    ws.append([f'Panchayath Name: {panchayat_name}'])
    ws.append([])
    ws.append(['Muster Roll No.', 'Image'])
    img_row = ws.max_row + 1
    img_refs = []
    for entry in image_records:
        muster_no = entry['muster_roll_no']
        img_bytes = entry['image']
        if img_bytes:
            img_bytes.seek(0)
            img_data = img_bytes.read()
            img_bytes_for_excel = io.BytesIO(img_data)
            xl_img = XLImage(img_bytes_for_excel)
            img_refs.append(img_bytes_for_excel)
            cell = ws.cell(row=img_row, column=1, value=muster_no)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=16, bold=True)
            cell_ref = f'B{img_row}'
            ws.add_image(xl_img, cell_ref)
            try:
                from PIL import Image as PILImage
                img_bytes_for_excel.seek(0)
                pil_img = PILImage.open(img_bytes_for_excel)
                img_height_px = pil_img.height
                img_height_pt = img_height_px * 0.75
                ws.row_dimensions[img_row].height = img_height_pt
                ws.column_dimensions['B'].width = 20
            except Exception:
                pass
        else:
            cell = ws.cell(row=img_row, column=1, value=muster_no)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=16, bold=True)
            ws.cell(row=img_row, column=2, value='No Image')
        img_row += 1
    print(f'Saved attendance_images_{file_base}.xlsx')
    return wb


def write_attendance_images_excel(option_c_records, work_code, work_name, panchayat_name, file_base):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance+Images'
    ws.append([f'Work Code: {work_code}'])
    ws.append([f'Work Name: {work_name if work_name else ""}'])
    ws.append([f'District: {DEFAULT_DISTRICT}'])
    ws.append([f'Taluk/Block: {DEFAULT_TALUK}'])
    ws.append([f'Panchayath Name: {panchayat_name}'])
    ws.append([])
    table_header = ['Muster Roll No.', 'S.No', 'Job Card No', 'Worker Name(Gender)', 'Attendance Date', 'Present/Absent', 'Image']
    ws.append(table_header)
    img_refs = []
    for entry in option_c_records:
        muster_no = entry['muster_roll_no']
        att_rows = entry['attendance']
        img_bytes = entry['image']
        first_row = True
        if att_rows:
            for row in att_rows:
                if len(row) >= 4 and row[3]:
                    row[3] = ' '.join(row[3].split()[:3])
                excel_row = [muster_no if first_row else '',
                             row[0] if len(row) > 0 else '',
                             row[1] if len(row) > 1 else '',
                             row[2] if len(row) > 2 else '',
                             row[3] if len(row) > 3 else '',
                             row[4] if len(row) > 4 else '',
                             '']
                if first_row:
                    if img_bytes:
                        img_bytes.seek(0)
                        img_data = img_bytes.read()
                        img_bytes_for_excel = io.BytesIO(img_data)
                        xl_img = XLImage(img_bytes_for_excel)
                        img_refs.append(img_bytes_for_excel)
                        row_idx = ws.max_row + 1
                        ws.append(excel_row)
                        cell_ref = f'G{row_idx}'
                        ws.add_image(xl_img, cell_ref)
                        ws.row_dimensions[row_idx].height = 100
                        ws.column_dimensions['G'].width = 20
                    else:
                        ws.append(excel_row)
                    first_row = False
                else:
                    ws.append(excel_row)
        else:
            excel_row = [muster_no, '', '', '', '', '', '']
            if img_bytes:
                img_bytes.seek(0)
                img_data = img_bytes.read()
                img_bytes_for_excel = io.BytesIO(img_data)
                xl_img = XLImage(img_bytes_for_excel)
                img_refs.append(img_bytes_for_excel)
                row_idx = ws.max_row + 1
                ws.append(excel_row)
                cell_ref = f'G{row_idx}'
                ws.add_image(xl_img, cell_ref)
                ws.row_dimensions[row_idx].height = 100
                ws.column_dimensions['G'].width = 20
            else:
                ws.append(excel_row)
    print(f'Saved attendance_with_images_{file_base}.xlsx')
    return wb


def run_attendance_downloader(panchayat_name, panchayat_code, fin_year, work_code, msr_start, msr_end, attendance_date, digest, progress_callback=None):
    attendance_records = []
    image_records = []
    option_c_records = []
    work_name = None
    table_headers = None
    for msr_no in range(msr_start, msr_end + 1):
        url_with_workcode = (
            f"{STARTING_URL}"
            f"panchayat_name={panchayat_name}&panchayat_code={panchayat_code}"
            f"&fin_year={fin_year}"
            f"&source=&work_code={work_code}"
            f"&msr_no={msr_no}"
            f"&AttendanceDate={attendance_date}"
            f"&Digest={digest}"
        )
        url_without_workcode = (
            f"{STARTING_URL}"
            f"panchayat_name={panchayat_name}&panchayat_code={panchayat_code}"
            f"&fin_year={fin_year}"
            f"&source="
            f"&msr_no={msr_no}"
            f"&AttendanceDate={attendance_date}"
            f"&Digest={digest}"
        )
        att_data, photo_url, wname, headers = get_attendance_data(url_with_workcode)
        if not att_data:
            att_data, photo_url, wname, headers = get_attendance_data(url_without_workcode)
        if wname and not work_name:
            work_name = wname
        if headers and not table_headers:
            table_headers = headers
        if att_data:
            for row in att_data:
                attendance_records.append({'muster_roll_no': msr_no, 'row': row})
        if photo_url:
            img_bytes = download_photo(photo_url)
            image_records.append({'muster_roll_no': msr_no, 'image': img_bytes})
        else:
            image_records.append({'muster_roll_no': msr_no, 'image': None})
        option_c_records.append({'muster_roll_no': msr_no, 'attendance': att_data, 'image': img_bytes if photo_url else None})
        print(f"Muster Roll No. {msr_no} parsed ")
        if progress_callback:
            progress_callback(f"Muster Roll {msr_no} parsed ,")
    file_base = f"{work_code}_{attendance_date}".replace('/', '_')

    att_wb = write_attendance_excel(attendance_records, work_code, work_name, panchayat_name, file_base)
    img_wb = write_images_excel(image_records, work_code, work_name, panchayat_name, file_base)
    optc_wb = write_attendance_images_excel(option_c_records, work_code, work_name, panchayat_name, file_base)

    att_xlsx = io.BytesIO()
    att_wb.save(att_xlsx)
    att_xlsx.seek(0)
    img_xlsx = io.BytesIO()
    img_wb.save(img_xlsx)
    img_xlsx.seek(0)
    optc_xlsx = io.BytesIO()
    optc_wb.save(optc_xlsx)
    optc_xlsx.seek(0)
    return att_xlsx, img_xlsx, optc_xlsx, None

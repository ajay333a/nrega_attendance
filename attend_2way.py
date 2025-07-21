import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import time
import openpyxl
from openpyxl.styles import Alignment, Font
import io
from attendance_downloader import get_attendance_data, download_photo
from openpyxl.drawing.image import Image as XLImage
import re
from concurrent.futures import ThreadPoolExecutor

# Constants
BASE_URL = "https://mnregaweb4.nic.in/nregaarch/View_NMMS_atten_date_new.aspx?fin_year=2024-2025&Digest=HNrisV4bhHnb7Gve3mAKYQ"
STATE_VALUE = '15'  # Karnataka
DISTRICT_NAME = 'BALLARI'
BLOCK_NAME = 'SIRUGUPPA'
TALUK_NAME = 'Siruguppa'
DISTRICT_LABEL = 'Ballari'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

def get_table_by_id_or_div(soup, table_id='grdTable', div_id='RepPr1'):
    table = soup.find('table', {'id': table_id})
    if not table:
        div = soup.find('div', {'id': div_id})
        if div:
            table = div.find('table')
    return table

def get_link_from_table(table, match_col_idx, match_text, href_col_idx=None):
    for row in table.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) > match_col_idx and cols[match_col_idx].get_text(strip=True).upper() == match_text.upper():
            a = cols[match_col_idx].find('a', href=True)
            if a:
                return a['href']
    return None

def get_panchayath_link(table, panchayath_name):
    for row in table.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) >= 4:
            s_no = cols[0].get_text(strip=True)
            if not s_no.isdigit():
                continue
            panch_name = cols[1].get_text(strip=True).upper()
            muster_rolls_a = cols[3].find('a', href=True)
            href = muster_rolls_a['href'] if muster_rolls_a else None
            if panch_name == panchayath_name and href:
                return href
    return None

def get_muster_roll_rows(muster_table, choice, workcode=None, workcode_idx=None, muster_no_idx=None):
    rows_to_save = []
    for row in muster_table.find_all('tr')[1:]:
        cols = row.find_all('td')
        if len(cols) > muster_no_idx:
            if choice == 'all':
                muster_a = cols[muster_no_idx].find('a', href=True)
                if muster_a:
                    muster_href = muster_a['href']
                    rows_to_save.append((cols, muster_href))
            elif choice == 'work' and workcode and workcode_idx is not None:
                if len(cols) > workcode_idx and workcode in cols[workcode_idx].get_text(strip=True):
                    muster_a = cols[muster_no_idx].find('a', href=True)
                    if muster_a:
                        muster_href = muster_a['href']
                        rows_to_save.append((cols, muster_href))
    return rows_to_save

def prompt_user_for_inputs(date_options):
    print("Available dates:", date_options)
    attendance_date = input("Enter attendance date from above options (e.g., 18/07/2025): ").strip()
    panchayath_name = input("Enter Panchayath name: ").strip().upper()
    choice = input("Type 'all' for all muster rolls or 'work' for specific work: ").strip().lower()
    workcode = None
    if choice == 'work':
        workcode = input("Enter workcode: ").strip()
    return attendance_date, panchayath_name, choice, workcode

def save_attendance_excel(wb, ws, img_wb, img_ws, panchayath_name, attendance_date):
    wb.save(f"muster_rolls_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")
    img_wb.save(f"muster_roll_images_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")
    print(f"Saved muster_rolls_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")
    print(f"Saved muster_roll_images_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")

def save_raw_excel(rows_to_save, panchayath_name, attendance_date, muster_no_idx, workcode_idx, panchayath_url, muster_data_cache):
    raw_wb = openpyxl.Workbook()
    raw_ws = raw_wb.active
    raw_ws.append([
        "Taluk", "Panchayath", "Work Code", "Muster Roll No", "Job Card No", "Worker Name", "Gender", "Attendance", "Attendance Date"
    ])
    for cols, muster_href in rows_to_save:
        muster_url = urljoin(panchayath_url, muster_href)
        attendance_data, _, _, _ = muster_data_cache.get(muster_url, (None, None, None, None))
        muster_roll_no = cols[muster_no_idx].get_text(strip=True)
        work_code = cols[workcode_idx].get_text(strip=True)
        for att_row in attendance_data or []:
            worker_name_full = att_row[2] if len(att_row) > 2 else ''
            if worker_name_full.endswith(')') and '(' in worker_name_full:
                name_part = worker_name_full[:worker_name_full.rfind('(')].strip()
                gender_part = worker_name_full[worker_name_full.rfind('(')+1:-1].strip()
            else:
                name_part = worker_name_full
                gender_part = ''
            raw_ws.append([
                TALUK_NAME,
                panchayath_name,
                work_code,
                muster_roll_no,
                att_row[1] if len(att_row) > 1 else '',
                name_part,
                gender_part,
                att_row[4] if len(att_row) > 4 else '',
                att_row[3] if len(att_row) > 3 else '',
            ])
    raw_wb.save(f"muster_rolls_raw_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")
    print(f"Saved muster_rolls_raw_{panchayath_name}_{attendance_date.replace('/', '_')}.xlsx")

def find_col_idx(header_cols, search):
    search_clean = re.sub(r'[^a-zA-Z0-9]', '', search.lower())
    for i, h in enumerate(header_cols):
        h_clean = re.sub(r'[^a-zA-Z0-9]', '', h.lower())
        if search_clean in h_clean:
            return i
    return None

def fetch_muster_data(muster_url):
    return get_attendance_data(muster_url)

def main():
    session = requests.Session()
    resp = session.get(BASE_URL, headers=HEADERS)
    soup = BeautifulSoup(resp.content, 'html.parser')

    viewstate = soup.find('input', {'id': '__VIEWSTATE'})['value']
    eventvalidation = soup.find('input', {'id': '__EVENTVALIDATION'})['value']
    viewstategen = soup.find('input', {'id': '__VIEWSTATEGENERATOR'})['value']

    attendance_select = soup.find('select', {'name': 'ctl00$ContentPlaceHolder1$ddl_attendance'})
    date_options = [opt['value'] for opt in attendance_select.find_all('option')]
    attendance_date, panchayath_name, choice, workcode = prompt_user_for_inputs(date_options)

    data = {
        '__VIEWSTATE': viewstate,
        '__VIEWSTATEGENERATOR': viewstategen,
        '__EVENTVALIDATION': eventvalidation,
        'ctl00$ContentPlaceHolder1$ddlstate': STATE_VALUE,
        'ctl00$ContentPlaceHolder1$ddl_attendance': attendance_date,
        'ctl00$ContentPlaceHolder1$btn_showreport': 'Show Attendance',
    }
    headers_post = HEADERS.copy()
    headers_post['Referer'] = BASE_URL
    resp2 = session.post(BASE_URL, data=data, headers=headers_post)
    # time.sleep(3)
    soup2 = BeautifulSoup(resp2.content, 'html.parser')

    # State table navigation
    state_table = get_table_by_id_or_div(soup2)
    if not state_table:
        print("Could not find state table.")
        return
    karnataka_link = get_link_from_table(state_table, 1, 'KARNATAKA')
    if not karnataka_link:
        print("Could not find Karnataka link in state table.")
        return
    karnataka_url = urljoin(BASE_URL, karnataka_link)

    # Districts table navigation
    resp3 = session.get(karnataka_url, headers=HEADERS)
    soup3 = BeautifulSoup(resp3.content, 'html.parser')
    dist_table = get_table_by_id_or_div(soup3)
    if not dist_table:
        print("Could not find districts table.")
        return
    ballari_link = get_link_from_table(dist_table, 1, DISTRICT_NAME)
    if not ballari_link:
        print("Could not find Ballari link in districts table.")
        return
    ballari_url = urljoin(karnataka_url, ballari_link)

    # Block/Taluk table navigation
    resp4 = session.get(ballari_url, headers=HEADERS)
    soup4 = BeautifulSoup(resp4.content, 'html.parser')
    block_table = get_table_by_id_or_div(soup4)
    if not block_table:
        print("Could not find block/taluk table.")
        return
    siruguppa_link = get_link_from_table(block_table, 1, BLOCK_NAME)
    if not siruguppa_link:
        print("Could not find Siruguppa link in block/taluk table.")
        return
    siruguppa_url = urljoin(ballari_url, siruguppa_link)

    # Panchayath table navigation
    resp5 = session.get(siruguppa_url, headers=HEADERS)
    soup5 = BeautifulSoup(resp5.content, 'html.parser')
    panch_div = soup5.find('div', {'id': 'RepPr1'})
    if not panch_div:
        print("Could not find panchayath table container.")
        return
    panch_table = panch_div.find('table')
    if not panch_table:
        print("Could not find panchayath table.")
        return
    panchayath_link = get_panchayath_link(panch_table, panchayath_name)
    if not panchayath_link:
        print("No NMR generated by the Panchayath")
        return
    panchayath_url = urljoin(siruguppa_url, panchayath_link)

    # Muster Roll table navigation
    resp6 = session.get(panchayath_url, headers=HEADERS)
    soup6 = BeautifulSoup(resp6.content, 'html.parser')
    muster_div = soup6.find('div', {'id': 'RepPr1'})
    if not muster_div:
        print("Could not find muster roll table container.")
        return
    muster_table = muster_div.find('table')
    if not muster_table:
        print("Could not find muster roll table.")
        return
    header_row = muster_table.find('tr')
    header_cols = [th.get_text(strip=True).replace('\u00a0', ' ').strip().lower() for th in header_row.find_all(['th', 'td'])]
    workcode_idx = find_col_idx(header_cols, 'work code')
    muster_no_idx = find_col_idx(header_cols, 'mustroll no')
    if workcode_idx is None or muster_no_idx is None:
        print("Could not find required columns in muster roll table header.")
        print("Header columns found:", header_cols)
        return

    rows_to_save = get_muster_roll_rows(muster_table, choice, workcode, workcode_idx, muster_no_idx)
    if not rows_to_save:
        print("No muster roll data found for the selection.")
        return

    # Excel setup
    wb = openpyxl.Workbook()
    ws = wb.active
    row_cursor = 1
    ws.cell(row=row_cursor, column=1, value="District:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=2, value=DISTRICT_LABEL)
    ws.cell(row=row_cursor, column=3, value="Taluk/Block:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=4, value=TALUK_NAME)
    row_cursor += 1
    ws.cell(row=row_cursor, column=1, value="Panchayath:").font = Font(bold=True)
    ws.cell(row=row_cursor, column=2, value=panchayath_name)
    row_cursor += 1
    workcode_row_idx_att = row_cursor
    ws.cell(row=workcode_row_idx_att, column=1, value="Work code:").font = Font(bold=True)
    ws.cell(row=workcode_row_idx_att, column=3, value="Work Name:").font = Font(bold=True)
    row_cursor += 1
    attendance_header_written = False
    first_muster_processed = False
    first_work_code = ''
    first_work_name = ''

    # Image-only Excel setup
    img_wb = openpyxl.Workbook()
    img_ws = img_wb.active
    img_row_cursor = 1
    img_bytes_refs = []
    img_ws.cell(row=img_row_cursor, column=1, value="District:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=2, value=DISTRICT_LABEL)
    img_ws.cell(row=img_row_cursor, column=3, value="Taluk/Block:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=4, value=TALUK_NAME)
    img_row_cursor += 1
    img_ws.cell(row=img_row_cursor, column=1, value="Panchayath:").font = Font(bold=True)
    img_ws.cell(row=img_row_cursor, column=2, value=panchayath_name)
    img_row_cursor += 1
    workcode_row_idx = img_row_cursor
    img_ws.cell(row=workcode_row_idx, column=1, value="Work code:").font = Font(bold=True)
    img_ws.cell(row=workcode_row_idx, column=3, value="Work Name:").font = Font(bold=True)
    img_row_cursor += 1
    header_row_idx = img_row_cursor
    img_ws.cell(row=header_row_idx, column=1, value='Muster Roll No').font = Font(bold=True)
    img_ws.cell(row=header_row_idx, column=2, value='Image').font = Font(bold=True)
    img_row_cursor += 1

    # Main loop: cache attendance data for each muster roll
    muster_data_cache = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        muster_urls = [urljoin(panchayath_url, href) for _, href in rows_to_save]
        results = list(executor.map(fetch_muster_data, muster_urls))
    # Now results[i] corresponds to muster_urls[i]

    for i, (muster_url, (attendance_data, photo_url, work_name, header_cells)) in enumerate(zip(muster_urls, results)):
        muster_data_cache[muster_url] = (attendance_data, photo_url, work_name, header_cells)
        img_bytes = download_photo(photo_url) if photo_url else None
        muster_roll_no = rows_to_save[i][0][muster_no_idx].get_text(strip=True)
        print(f"Muster Roll No. {muster_roll_no} parsed ")
        # Attendance Excel
        if not attendance_header_written and header_cells:
            if not first_muster_processed:
                first_work_code = rows_to_save[i][0][workcode_idx].get_text(strip=True)
                first_work_name = work_name or ''
                ws.cell(row=workcode_row_idx_att, column=2, value=first_work_code)
                ws.cell(row=workcode_row_idx_att, column=4, value=first_work_name)
                first_muster_processed = True
            ws.cell(row=row_cursor, column=1, value="Muster Roll No").font = Font(bold=True)
            for col_idx, header in enumerate(header_cells, 2):
                ws.cell(row=row_cursor, column=col_idx, value=header).font = Font(bold=True)
            row_cursor += 1
            attendance_header_written = True
        if attendance_data:
            for att_row in attendance_data:
                ws.cell(row=row_cursor, column=1, value=muster_roll_no)
                for col_idx, val in enumerate(att_row, 2):
                    ws.cell(row=row_cursor, column=col_idx, value=val)
                row_cursor += 1
        if img_bytes:
            img_bytes.seek(0)
            img = XLImage(img_bytes)
            img_cell = f"H{row_cursor-len(attendance_data) if attendance_data else row_cursor}"
            ws.add_image(img, img_cell)
            row_cursor += 3
        else:
            row_cursor += 2
        row_cursor += 2
        # Image-only Excel
        start_img_row = img_row_cursor
        img_ws.cell(row=img_row_cursor, column=1, value=muster_roll_no).font = Font(bold=True, size=18)
        if img_bytes:
            img_bytes.seek(0)
            img_bytes_for_imgwb = io.BytesIO(img_bytes.getbuffer())
            img2 = XLImage(img_bytes_for_imgwb)
            img_ws.add_image(img2, f"B{img_row_cursor}")
            img_bytes_refs.append(img_bytes_for_imgwb)
            img_height_rows = 20
            end_img_row = img_row_cursor + img_height_rows - 1
            img_ws.merge_cells(start_row=start_img_row, start_column=1, end_row=end_img_row, end_column=1)
            img_ws.cell(row=start_img_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
            img_row_cursor += img_height_rows
        else:
            end_img_row = img_row_cursor
            img_row_cursor += 3
            img_ws.merge_cells(start_row=start_img_row, start_column=1, end_row=end_img_row, end_column=1)
            img_ws.cell(row=start_img_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
        img_row_cursor += 2
    if not first_muster_processed:
        img_ws.cell(row=workcode_row_idx, column=2, value='')
        img_ws.cell(row=workcode_row_idx, column=4, value='')
    save_attendance_excel(wb, ws, img_wb, img_ws, panchayath_name, attendance_date)
    save_raw_excel(rows_to_save, panchayath_name, attendance_date, muster_no_idx, workcode_idx, panchayath_url, muster_data_cache)

if __name__ == "__main__":
    main()
